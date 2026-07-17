"""Render a :class:`ReportDocument` to an .xlsx workbook via openpyxl."""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from reports.report_tables import ReportDocument, ReportTable

_NUMBER_FORMAT = "#,##0;(#,##0)"
_HEADER_FILL = "FF1F2937"
_TOTAL_FILL = "FFEFF3F8"
_TITLE_SIZE = 14


def export_excel(doc: ReportDocument, path: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=_TITLE_SIZE)
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    total_fill = PatternFill("solid", fgColor=_TOTAL_FILL)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    max_cols = max((len(t.columns) for t in doc.tables), default=1)
    row = 1

    title_cell = ws.cell(row=row, column=1, value=doc.title)
    title_cell.font = title_font
    title_cell.alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
    row += 1
    sub_cell = ws.cell(row=row, column=1, value=doc.subtitle)
    sub_cell.alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
    row += 2

    col_widths: dict[int, int] = {}

    def _width(col: int, text: str) -> None:
        col_widths[col] = max(col_widths.get(col, 10), min(len(text) + 2, 48))

    for table in doc.tables:
        if table.caption:
            cap = ws.cell(row=row, column=1, value=table.caption)
            cap.font = bold
            row += 1

        for c, column in enumerate(table.columns, start=1):
            cell = ws.cell(row=row, column=c, value=column.header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = right if column.numeric else center
            _width(c, column.header)
        row += 1

        for data_row in table.rows:
            _write_row(ws, row, table, data_row, right, _NUMBER_FORMAT, _width)
            row += 1

        if table.total_row is not None:
            _write_row(ws, row, table, table.total_row, right, _NUMBER_FORMAT, _width)
            for c in range(1, len(table.columns) + 1):
                ws.cell(row=row, column=c).font = bold
                ws.cell(row=row, column=c).fill = total_fill
            row += 1
        row += 1  # blank spacer between tables

    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    path = Path(path)
    wb.save(path)
    return path


def _write_row(ws, row, table: ReportTable, values, right, num_fmt, width_fn) -> None:
    for c, (column, value) in enumerate(zip(table.columns, values), start=1):
        cell = ws.cell(row=row, column=c)
        if isinstance(value, Decimal):
            cell.value = float(value)
            cell.number_format = num_fmt
            cell.alignment = right
            width_fn(c, f"{value:,.0f}")
        elif value is None:
            cell.value = None
        else:
            cell.value = str(value)
            width_fn(c, str(value))
